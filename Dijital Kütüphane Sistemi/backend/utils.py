"""
Yardımcı fonksiyonlar ve validasyon araçları
"""
import re
import os
from datetime import datetime, date
from typing import Optional, Tuple, List
import hashlib


class Validator:
    """Veri doğrulama sınıfı"""
    
    @staticmethod
    def email(email: str) -> Tuple[bool, str]:
        """Email doğrulama"""
        if not email or not email.strip():
            return False, "Email adresi boş olamaz"
        
        email = email.strip().lower()
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        
        if not re.match(pattern, email):
            return False, "Geçersiz email formatı. Örnek: kullanici@domain.com"
        
        return True, "Geçerli"
    
    @staticmethod
    def telefon(telefon: str) -> Tuple[bool, str]:
        """Telefon numarası doğrulama"""
        if not telefon:
            return True, ""  # Telefon opsiyonel
        
        temiz = re.sub(r'[\s\-\(\)]+', '', telefon)
        pattern = r'^(\+90|0)?[5][0-9]{9}$'
        
        if not re.match(pattern, temiz):
            return False, "Geçersiz telefon. Örnek: 05XX XXX XX XX"
        
        return True, "Geçerli"
    
    @staticmethod
    def isbn(isbn: str) -> Tuple[bool, str]:
        """ISBN doğrulama"""
        if not isbn:
            return True, ""  # ISBN opsiyonel
        
        temiz = re.sub(r'[\s\-]+', '', isbn)
        
        if len(temiz) == 10:
            # ISBN-10 kontrolü
            try:
                toplam = sum((i + 1) * int(d) for i, d in enumerate(temiz[:-1]))
                kontrol = toplam % 11
                kontrol_char = 'X' if kontrol == 10 else str(kontrol)
                if temiz[-1].upper() == kontrol_char:
                    return True, "Geçerli ISBN-10"
            except:
                pass
        
        elif len(temiz) == 13:
            # ISBN-13 kontrolü
            try:
                toplam = sum((1 if i % 2 == 0 else 3) * int(d) for i, d in enumerate(temiz[:-1]))
                kontrol = (10 - (toplam % 10)) % 10
                if int(temiz[-1]) == kontrol:
                    return True, "Geçerli ISBN-13"
            except:
                pass
        
        return False, "Geçersiz ISBN formatı"
    
    @staticmethod
    def sifre_gucu(sifre: str) -> Tuple[bool, str, int]:
        """Şifre gücü kontrolü"""
        if len(sifre) < 8:
            return False, "Şifre en az 8 karakter olmalı", 0
        
        puan = 0
        
        if len(sifre) >= 12: puan += 1
        if re.search(r'[A-Z]', sifre): puan += 1
        if re.search(r'[a-z]', sifre): puan += 1
        if re.search(r'\d', sifre): puan += 1
        if re.search(r'[!@#$%^&*(),.?":{}|<>]', sifre): puan += 1
        
        gucler = {0: "Çok Zayıf", 1: "Zayıf", 2: "Orta", 3: "Güçlü", 4: "Çok Güçlü", 5: "Mükemmel"}
        
        if puan < 2:
            return False, f"Şifre çok zayıf ({gucler[puan]})", puan
        
        return True, f"Şifre gücü: {gucler[puan]}", puan
    
    @staticmethod
    def bos_mu(text: str) -> bool:
        """Metin boş mu kontrolü"""
        return not text or not text.strip()
    
    @staticmethod
    def pozitif_sayi_mi(deger) -> bool:
        """Pozitif sayı kontrolü"""
        try:
            return float(deger) > 0
        except:
            return False


class DateHelper:
    """Tarih işlemleri yardımcı sınıfı"""
    
    @staticmethod
    def parse_date(tarih_str: str) -> Optional[date]:
        """String'den date objesine dönüştür"""
        formats = ['%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y/%m/%d']
        
        for fmt in formats:
            try:
                return datetime.strptime(tarih_str, fmt).date()
            except:
                continue
        
        return None
    
    @staticmethod
    def format_date(tarih: date, format: str = '%d.%m.%Y') -> str:
        """Date objesini string'e dönüştür"""
        return tarih.strftime(format)
    
    @staticmethod
    def bugun() -> date:
        return date.today()
    
    @staticmethod
    def ay_basi() -> date:
        bugun = date.today()
        return date(bugun.year, bugun.month, 1)
    
    @staticmethod
    def yil_basi() -> date:
        return date(date.today().year, 1, 1)


class ExportManager:
    """Dışa aktarma işlemleri"""
    
    @staticmethod
    def _ensure_export_dirs():
        """Export klasörlerini oluştur"""
        os.makedirs("exports/csv", exist_ok=True)
        os.makedirs("exports/pdf", exist_ok=True)
        os.makedirs("exports/excel", exist_ok=True)
    
    @staticmethod
    def to_csv(data: List[dict], filename: str, columns: List[str] = None) -> str:
        """CSV'ye aktar"""
        import csv
        
        ExportManager._ensure_export_dirs()
        
        filepath = f"exports/csv/{filename}"
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            if not columns and data:
                columns = list(data[0].keys())
            
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            writer.writerows(data)
        
        print(f"✓ CSV oluşturuldu: {filepath}")
        return filepath
    
    @staticmethod
    def to_pdf(data: List[dict], title: str, filename: str) -> str:
        """PDF'e aktar"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import mm
        except ImportError:
            print("⚠️ reportlab yüklü değil. pip install reportlab")
            return ""
        
        ExportManager._ensure_export_dirs()
        
        filepath = f"exports/pdf/{filename}"
        
        c = canvas.Canvas(filepath, pagesize=A4)
        width, height = A4
        
        # Başlık
        c.setFont("Helvetica-Bold", 16)
        c.drawString(20*mm, height - 20*mm, title)
        
        # Tarih
        c.setFont("Helvetica", 10)
        c.drawString(20*mm, height - 25*mm, f"Oluşturma: {date.today().strftime('%d.%m.%Y')}")
        
        # Veriler
        c.setFont("Helvetica", 8)
        y = height - 35*mm
        
        for row in data[:50]:  # İlk 50 satır
            text = " | ".join(str(v) for v in row.values())
            c.drawString(20*mm, y, text[:100])  # Max 100 karakter
            y -= 5*mm
            
            if y < 20*mm:
                c.showPage()
                y = height - 20*mm
        
        c.save()
        print(f"✓ PDF oluşturuldu: {filepath}")
        return filepath
    
    @staticmethod
    def to_excel(data: List[dict], filename: str, sheet_name: str = "Veri") -> str:
        """Excel'e aktar"""
        try:
            import openpyxl
        except ImportError:
            print("⚠️ openpyxl yüklü değil. pip install openpyxl")
            return ""
        
        ExportManager._ensure_export_dirs()
        
        filepath = f"exports/excel/{filename}"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if data:
            # Başlıklar
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Veriler
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, key in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data[key])
        
        wb.save(filepath)
        print(f"✓ Excel oluşturuldu: {filepath}")
        return filepath


class FileManager:
    """Dosya yönetimi"""
    
    @staticmethod
    def ensure_directories():
        """Tüm gerekli klasörleri oluştur"""
        dirs = [
            "data",
            "logs", 
            "exports",
            "exports/csv",
            "exports/pdf", 
            "exports/excel",
            "backups"
        ]
        
        created = []
        for d in dirs:
            if not os.path.exists(d):
                os.makedirs(d)
                created.append(d)
        
        return created
    
    @staticmethod
    def get_file_size(filepath: str) -> str:
        """Dosya boyutunu okunabilir formatta döndür"""
        if not os.path.exists(filepath):
            return "0 B"
        
        size = os.path.getsize(filepath)
        
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size < 1024:
                return f"{size:.2f} {unit}"
            size /= 1024
        
        return f"{size:.2f} TB"
    
    @staticmethod
    def clean_exports(max_age_days: int = 7):
        """Eski export dosyalarını temizle"""
        from datetime import timedelta
        
        cutoff = datetime.now() - timedelta(days=max_age_days)
        deleted = 0
        
        for root, dirs, files in os.walk("exports"):
            for file in files:
                filepath = os.path.join(root, file)
                if datetime.fromtimestamp(os.path.getmtime(filepath)) < cutoff:
                    os.remove(filepath)
                    deleted += 1
        
        return deleted